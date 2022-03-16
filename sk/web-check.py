import csv
from os.path import join
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

HTML_TAGS = {
    "a",
    "abbr",
    "address",
    "area",
    "article",
    "aside",
    "audio",
    "b",
    "base",
    "bdi",
    "bdo",
    "blockquote",
    "body",
    "br",
    "button",
    "canvas",
    "caption",
    "cite",
    "code",
    "col",
    "colgroup",
    "datalist",
    "dd",
    "del",
    "details",
    "dfn",
    "dialog",
    "div",
    "dl",
    "dt",
    "em",
    "embed",
    "fieldset",
    "figcaption",
    "figure",
    "footer",
    "form",
    "h1",
    "head",
    "header",
    "hr",
    "html",
    "i",
    "iframe",
    "img",
    "input",
    "ins",
    "kbd",
    "label",
    "legend",
    "li",
    "link",
    "main",
    "map",
    "mark",
    "menu",
    "menuitem",
    "meta",
    "meter",
    "nav",
    "noscript",
    "object",
    "ol",
    "optgroup",
    "option",
    "output",
    "p",
    "param",
    "picture",
    "pre",
    "progress",
    "q",
    "rp",
    "rt",
    "ruby",
    "s",
    "samp",
    "script",
    "section",
    "select",
    "small",
    "source",
    "span",
    "strong",
    "style",
    "sub",
    "summary",
    "sup",
    "table",
    "tbody",
    "td",
    "textarea",
    "tfoot",
    "th",
    "thead",
    "time",
    "title",
    "tr",
    "track",
    "ul",
    "var",
    "video",
    "wbr",
    "acronym",
    "applet",
    "basefont",
    "big",
    "center",
    "dir",
    "font",
    "frame",
    "frameset",
    "noframes",
    "strike",
    "tt",
    "u",
}


OTHERS_TAGS = HTML_TAGS - {
    'html',
    'head',
    'meta',
    'title',
    'style',
    'script',
    'body',
    'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
    'p',
    'a',
    'img',
    'ul',
    'ol',
    'li',
    'em',
    'strong',
    'i',
    'u',
    'b',
    'div',
    'span',
}


def read_student(csv_filename):
    with open(csv_filename, encoding="utf-8") as f:
        for row in csv.reader(f, delimiter=','):
            yield row


def download_html(src_url, dst_dir, student):
    Path(dst_dir).mkdir(exist_ok=True, parents=True)

    if not url.strip():
        return '', BeautifulSoup('', 'html.parser')

    std_html = requests.get(url)

    html_path = join(dst_dir, f"{student}.html")
    with open(html_path, "w", encoding="utf8") as f:
        try:
            html_text = std_html.content.decode('utf-8')
        except UnicodeDecodeError:
            html_text = std_html.content.decode('windows-1251')
        f.write(html_text)
    return html_path, BeautifulSoup(html_text, 'html.parser')


def check_python():
    return [
        ("Python", ""),
        ("В коде есть переменные разных типов (bool, int, float, str)", ''),
        ("В коде есть операторы ветвления (if/elif/else)", ''),
        ("В коде есть операторы циклов (while/for)", ''),
        ("В коде есть функции (def ...())", ''),
        ("Вы водится таблица с 5-ю столбцами (№ Месяца, Задолженность, Мин. Платеж, Процент, Платеж в банк)", ''),
        ("В конце выводиться кол-во месяцев", ''),
        ("В конце выводиться кол-во заплаченных денег", ''),
        ("В конце выводиться кол-во денег, заплаченных как проценты", ''),
        ("Программа форматирует вывод (бонус)", ''),
        ("Самсостоятельные, творческие решения (бонус)", ''),
    ]


def check_html(url, bs_text):
    if not url.strip():
        return [
            ("HTML & CSS", ""),
            ("Полноценная структура HTML-документа", ''),
            ("В HTML-коде присутствуют Заголовки (h1, h2, ..., h6)", ''),
            ("В HTML-коде присутствуют Абзацы, Параграфы (p)", ''),
            ("В HTML-коде присутствуют Изображения", ''),
            ("В HTML-коде присутствуют Списки", ''),
            ("В HTML-коде присутствуют Ссылки", ''),
            ("В HTML-коде присутствуют Прочие HTML-теги (+1 за тег)", ''),
            ("Есть CSS-правила в теге style", ''),
            ("Подключен и используется Bootstrap", ''),
            ("Эстетичный вид HTML-страницы", ''),
        ]

    html = bs_text.find('html')
    head = html and html.find('head')
    title = head and head.find('title')
    meta_charset = head and head.find('meta', {'charset': True})
    body = html and html.find('body')

    ul = body and body.find('ul')
    ol = body and body.find('ol')
    li = (ul and ul.find('li')) or (ol and ol.find('li'))

    used_other_tags = sum([int(bool(html and html.find(tag)))
                          for tag in OTHERS_TAGS])

    return [
        ("HTML & CSS", ""),
        ("Полноценная структура HTML-документа",
            int(bool(html and head and title and meta_charset and body))),
        ("В HTML-коде присутствуют Заголовки (h1, h2, ..., h6)",
            int(bool(html and html.find(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])))),
        ("В HTML-коде присутствуют Абзацы, Параграфы (p)",
            int(bool(body and body.find('p')))),
        ("В HTML-коде присутствуют Изображения",
            int(bool(body and body.find('img', {'src': True})))),
        ("В HTML-коде присутствуют Списки",
            int(bool(li))),
        ("В HTML-коде присутствуют Ссылки",
            int(bool(body and body.find('a', {'href': True})))),
        ("В HTML-коде присутствуют Прочие HTML-теги (+1 за тег)",
            min(3, used_other_tags)),
        ("Есть CSS-правила в теге style",
            int(bool(html and html.find('style')))),
        ("Подключен и используется Bootstrap", ''),
        ("Эстетичный вид HTML-страницы", ''),
    ]


def check_web(url, bs_text):
    if not url.strip():
        return [
            ("WEB", ""),
            ("Страница выложена на GitHub pages", ''),
            ("Все картинки на страницы загружаются", ''),
            ("Ссылки на странице работают", ''),
        ]

    img_loaded = False
    for img in bs_text.find_all('img'):
        if not img.has_attr('src'):
            img_loaded = False
            break

        img['src'] = img['src'].strip()

        if img['src'].startswith('data:'):
            continue

        try:
            if img['src'].startswith('http'):
                    res = requests.get(img['src'], timeout=1)
            else:
                res = requests.get(f'{url}/{img["src"]}', timeout=1)
            if res.status_code != 200:
                img_loaded = False
                break
        except Exception as e:
            print(e)
    else:
        img_loaded = bool(len(bs_text.find_all('img')))

    hr_works = False
    for a in bs_text.find_all('a'):
        if not a.has_attr('href'):
            hr_works = False
            break

        a['href'] = a['href'].strip()
        try:
            if a['href'].startswith('http'):
                res = requests.get(a['href'], timeout=1)
            else:
                res = requests.get(f'{url}/{a["href"]}', timeout=1)
            if res.status_code != 200:
                hr_works = False
                break
        except Exception as e:
            print(e)

    else:
        hr_works = bool(len(bs_text.find_all('a')))

    return [
        ("WEB", ""),
        ("Страница выложена на GitHub pages", int("github.io" in url)),
        ("Все картинки на страницы загружаются", int(img_loaded)),
        ("Ссылки на странице работают", int(hr_works)),
    ]


COLUMN = 0


def print_results(ws, group, student, *results,):
    global COLUMN
    if COLUMN == 0:
        ws['A2'] = 'Критерий'
        for i in range(len(results)):
            ws[f'A{i+3}'] = results[i][0]
    COLUMN += 1
    c = chr(ord("A") + COLUMN)
    if ord(c) > ord("Z"):
        return
    ws[f'{c}1'] = group
    ws[f'{c}2'] = student
    for i in range(len(results)):
        ws[f'{c}{i+3}'] = results[i][1]


if __name__ == "__main__":
    group = '114'
    group_html_dir = f'{group}/html'

    wb = Workbook()
    ws = wb.active

    for (student, url) in read_student(f'{group}.txt'):
        _, bs_text = download_html(url, group_html_dir, student)
        py_results = check_python()
        html_results = check_html(url, bs_text)
        web_results = check_web(url, bs_text)

        print_results(ws, group, student,
                      *py_results,
                      *html_results,
                      *web_results)
        wb.save(f"{group}.xlsx")
